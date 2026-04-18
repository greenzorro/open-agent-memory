"""
File: spreadsheet.py
Project: routine
Created: 2025-09-22 10:44:02
Author: Victor Cheng
Email: your_email@example.com
Description: 
"""

import os
import pandas as pd
from typing import Optional, Union, List
from .basic import *


# ==================== 内部辅助函数 ====================

def _is_engine_available(engine_name: str) -> bool:
    """检查指定的Excel引擎是否可用"""
    try:
        if engine_name == 'xlrd':
            import xlrd
            return True
        elif engine_name == 'xlwt':
            import xlwt
            return True
        elif engine_name == 'openpyxl':
            import openpyxl
            return True
        else:
            # 对于其他引擎，尝试导入
            try:
                __import__(engine_name)
                return True
            except ImportError:
                return False
    except ImportError:
        return False


def _is_excel_file(file_path: str) -> bool:
    """检查文件是否为Excel文件格式"""
    return file_path.lower().endswith(('.xlsx', '.xls'))


def _detect_string_columns(df_preview: pd.DataFrame) -> set:
    """检测可能需要保持为字符串类型的列"""
    string_columns = set()
    for col in df_preview.columns:
        col_data = df_preview[col].dropna()
        if len(col_data) > 0:
            # 检查是否所有非空值看起来都像数字但可能是字符串
            all_look_like_strings = False
            if col_data.dtype in ['int64', 'float64']:
                # 检查是否所有值都是整数且在一定范围内（可能是字符串数字）
                try:
                    # 尝试将值转换回字符串，看是否保持原样
                    original_as_str = [str(x) for x in col_data]
                    # 如果所有字符串表示都只有数字，可能是字符串数字
                    if all(x.isdigit() for x in original_as_str):
                        # 进一步检查：如果值的范围合理，可能是字符串数字
                        if all(0 <= x <= 999999 for x in col_data):  # 合理的字符串数字范围
                            all_look_like_strings = True
                except:
                    pass
            
            if all_look_like_strings:
                string_columns.add(col)
    return string_columns


def _get_actual_sheet_name(file_path: str, sheet_name: Union[str, int]) -> str:
    """获取实际的工作表名称（如果是索引，转换为名称）"""
    if isinstance(sheet_name, int):
        # 获取工作表名称列表
        excel_file = pd.ExcelFile(file_path)
        return excel_file.sheet_names[sheet_name]
    else:
        return sheet_name


def _is_html_table_file(file_path: str) -> bool:
    """检查文件是否是HTML格式的表格"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read(1024)  # 只读取前1KB内容
            return '<html' in content.lower() and '<table' in content.lower()
    except:
        return False


def _convert_html_table_to_csv(html_path: str, csv_path: str) -> bool:
    """将HTML表格转换为CSV文件

    复用basic.py中的html_table_2_csv_content函数，保持输出格式完全一致。
    basic.py返回的格式：list[list]，每行一个列表，保留所有原始行。
    """
    try:
        print(f"转换HTML表格: {html_path} -> {csv_path}")

        # 读取HTML内容
        with open(html_path, 'r', encoding='utf-8') as f:
            html_content = f.read()

        # 使用basic.py中的现成函数
        csv_data = html_table_2_csv_content(html_content)

        if csv_data:
            # 转换为DataFrame并保存（header=False保持与basic.py输出一致）
            df = pd.DataFrame(csv_data)
            df.to_csv(csv_path, index=False, header=False, encoding='utf-8-sig')
            print(f"成功转换HTML表格: {html_path} -> {csv_path}")
            print(f"数据形状: {df.shape}")
            return True
        else:
            print("错误: 未找到有效数据")
            return False

    except Exception as e:
        print(f"转换HTML表格失败: {e}")
        return False


# ==================== 公共函数 ====================


def get_excel_info(file_path: str) -> dict:
    """
    获取Excel文件的基本信息
    
    Args:
        file_path: Excel文件路径
    
    Returns:
        包含文件信息的字典
    
    Example:
        >>> info = get_excel_info('data.xlsx')
        >>> print(info)
        {
            'sheets': ['Sheet1', 'Sheet2'],
            'sheet_info': {
                'Sheet1': {
                    'rows': 100,
                    'columns': 5,
                    'column_names': ['日期', '销售额', '部门', '产品', '备注']
                }
            }
        }
    """
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    try:
        # 读取所有工作表
        excel_file = pd.ExcelFile(file_path)
        
        info = {
            'sheets': excel_file.sheet_names,
            'sheet_info': {}
        }
        
        # 获取每个工作表的详细信息
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            info['sheet_info'][sheet_name] = {
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': df.columns.tolist()
            }
        
        return info
        
    except Exception as e:
        raise Exception(f"获取Excel文件信息时出错: {str(e)}")


def sort_excel_by_column(
    file_path: str,
    sheet_name: Union[str, int],
    column_name: str,
    ascending: bool = True,
    output_path: Optional[str] = None
) -> str:
    """
    在Excel文件中根据指定列对数据进行排序
    
    Args:
        file_path: Excel文件路径 (.xls 或 .xlsx)
        sheet_name: 工作表名称或索引 (从0开始)
        column_name: 用于排序的列名 (第一行作为列名)
        ascending: 是否升序排列 (True=升序, False=降序)
        output_path: 输出文件路径 (可选，默认为原文件名加"_sorted")
    
    Returns:
        排序后的文件路径
        
    Raises:
        FileNotFoundError: 文件不存在
        ValueError: 工作表或列不存在，或文件格式不支持
        Exception: 其他处理错误
    
    Example:
        >>> sort_excel_by_column('data.xlsx', 'Sheet1', '日期', ascending=False)
        'data_sorted.xlsx'
        
        >>> sort_excel_by_column('report.xls', 0, '销售额', ascending=True)
        'report_sorted.xls'
    """
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    # 检查文件格式
    if not _is_excel_file(file_path):
        raise ValueError(f"不支持的文件格式。只支持 .xls 和 .xlsx 格式: {file_path}")
    
    try:
        # 读取Excel文件，智能检测需要保持字符串类型的列
        # 首先读取文件结构，检测可能的数字字符串列
        # 对于.xls文件，需要指定engine='xlrd'，对于.xlsx文件使用默认引擎
        if file_path.lower().endswith('.xls'):
            df_preview = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100, engine='xlrd')
        else:
            df_preview = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100)
        
        # 检测可能需要保持为字符串类型的列
        string_columns = _detect_string_columns(df_preview)
        
        # 使用检测到的数据类型重新读取完整数据
        # 对于可能是字符串数字的列，使用字符串类型读取
        dtype_dict = {col: str for col in string_columns}
        
        # 根据文件格式选择合适的读取引擎
        if file_path.lower().endswith('.xls'):
            # 对于.xls文件，优先使用xlrd引擎
            if _is_engine_available('xlrd'):
                df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=dtype_dict, engine='xlrd')
            else:
                raise ImportError("处理.xls文件需要xlrd引擎，请安装: pip install xlrd")
        else:
            # 对于.xlsx文件，使用默认引擎
            df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=dtype_dict)
        
        # 获取实际的工作表名称
        actual_sheet_name = _get_actual_sheet_name(file_path, sheet_name)
        
        # 检查列是否存在
        if column_name not in df.columns:
            available_columns = ', '.join(df.columns.tolist())
            raise ValueError(f"列 '{column_name}' 不存在。可用列: {available_columns}")
        
        # 对数据进行排序
        df_sorted = df.sort_values(by=column_name, ascending=ascending)
        
        # 生成输出文件路径
        if output_path is None:
            base_dir = os.path.dirname(file_path)
            base_name = os.path.basename(file_path)
            name_without_ext, ext = os.path.splitext(base_name)
            output_path = os.path.join(base_dir, f"{name_without_ext}_sorted{ext}")
        
        # 保存排序后的文件，保持原始数据类型
        # 根据文件扩展名选择保存方式
        if output_path.lower().endswith('.xlsx'):
            # 对于.xlsx文件，使用默认引擎
            # 对于字符串列，确保它们保持字符串格式
            for col in df_sorted.columns:
                if col in dtype_dict and dtype_dict[col] == str:
                    df_sorted[col] = df_sorted[col].astype(str)
            df_sorted.to_excel(output_path, sheet_name=actual_sheet_name, index=False)
        else:  # .xls
            # 对于.xls文件，优先使用xlwt引擎
            if _is_engine_available('xlwt'):
                # 对于字符串列，确保它们保持字符串格式
                for col in df_sorted.columns:
                    if col in dtype_dict and dtype_dict[col] == str:
                        df_sorted[col] = df_sorted[col].astype(str)
                df_sorted.to_excel(output_path, sheet_name=actual_sheet_name, index=False, engine='xlwt')
            else:
                # 如果xlwt不可用，提供明确的错误信息
                raise ImportError("保存.xls文件需要xlwt引擎，请安装: pip install xlwt")
        
        return output_path
        
    except FileNotFoundError:
        raise
    except ValueError as e:
        raise ValueError(f"数据验证错误: {str(e)}")
    except Exception as e:
        raise Exception(f"处理Excel文件时出错: {str(e)}")


def sort_excel_by_multiple_columns(
    file_path: str,
    sheet_name: Union[str, int],
    column_orders: List[tuple],
    output_path: Optional[str] = None
) -> str:
    """
    在Excel文件中根据多个列对数据进行排序
    
    Args:
        file_path: Excel文件路径 (.xls 或 .xlsx)
        sheet_name: 工作表名称或索引 (从0开始)
        column_orders: 排序规则列表，每个元素为 (列名, 是否升序) 的元组
                      例如: [('日期', False), ('销售额', True)] 表示先按日期降序，再按销售额升序
        output_path: 输出文件路径 (可选，默认为原文件名加"_sorted")
    
    Returns:
        排序后的文件路径
        
    Example:
        >>> sort_excel_by_multiple_columns(
        ...     'data.xlsx', 'Sheet1', 
        ...     [('部门', True), ('销售额', False)]
        ... )
        'data_sorted.xlsx'
    """
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    # 检查文件格式
    if not _is_excel_file(file_path):
        raise ValueError(f"不支持的文件格式。只支持 .xls 和 .xlsx 格式: {file_path}")
    
    try:
        # 读取Excel文件，智能检测需要保持字符串类型的列
        # 首先读取文件结构，检测可能的数字字符串列
        df_preview = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100)
        
        # 检测可能需要保持为字符串类型的列
        string_columns = _detect_string_columns(df_preview)
        
        # 使用检测到的数据类型重新读取完整数据
        # 对于可能是字符串数字的列，使用字符串类型读取
        dtype_dict = {col: str for col in string_columns}
        df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=dtype_dict)
        
        # 获取实际的工作表名称
        actual_sheet_name = _get_actual_sheet_name(file_path, sheet_name)
        
        # 检查所有列是否存在
        missing_columns = []
        for column_name, _ in column_orders:
            if column_name not in df.columns:
                missing_columns.append(column_name)
        
        if missing_columns:
            available_columns = ', '.join(df.columns.tolist())
            raise ValueError(f"以下列不存在: {', '.join(missing_columns)}。可用列: {available_columns}")
        
        # 准备排序参数
        sort_columns = [col for col, _ in column_orders]
        ascending_flags = [asc for _, asc in column_orders]
        
        # 对数据进行排序
        df_sorted = df.sort_values(by=sort_columns, ascending=ascending_flags)
        
        # 生成输出文件路径
        if output_path is None:
            base_dir = os.path.dirname(file_path)
            base_name = os.path.basename(file_path)
            name_without_ext, ext = os.path.splitext(base_name)
            output_path = os.path.join(base_dir, f"{name_without_ext}_sorted{ext}")
        
        # 保存排序后的文件
        # 根据文件扩展名选择保存方式
        if output_path.lower().endswith('.xlsx'):
            df_sorted.to_excel(output_path, sheet_name=actual_sheet_name, index=False)
        else:  # .xls
            try:
                df_sorted.to_excel(output_path, sheet_name=actual_sheet_name, index=False, engine='xlwt')
            except ValueError:
                # 如果xlwt不可用，使用openpyxl作为备选
                df_sorted.to_excel(output_path, sheet_name=actual_sheet_name, index=False, engine='openpyxl')
        
        return output_path
        
    except FileNotFoundError:
        raise
    except ValueError as e:
        raise ValueError(f"数据验证错误: {str(e)}")
    except Exception as e:
        raise Exception(f"处理Excel文件时出错: {str(e)}")


def get_latest_excel_file(directory: str) -> Optional[str]:
    """
    获取目录中最新的Excel文件（包括无扩展名的文件）

    Args:
        directory: 目录路径

    Returns:
        Optional[str]: 最新文件路径，如未找到则返回None
    """
    import os
    
    if not os.path.exists(directory):
        return None

    files = []
    for filename in os.listdir(directory):
        filepath = os.path.join(directory, filename)
        if os.path.isfile(filepath):
            # 检查是否为Excel文件
            if (filename.lower().endswith('.xls') or
                filename.lower().endswith('.xlsx') or
                'excel' in filename.lower() or
                # 检查文件内容（对于无扩展名的文件）
                is_likely_excel_file(filepath)):
                files.append((filepath, os.path.getmtime(filepath)))

    if not files:
        return None

    # 按修改时间排序，返回最新的文件
    files.sort(key=lambda x: x[1], reverse=True)
    return files[0][0]


def is_likely_excel_file(filepath: str) -> bool:
    """
    检查文件是否可能是Excel文件

    Args:
        filepath: 文件路径

    Returns:
        bool: 是否可能是Excel文件
    """
    import os
    
    try:
        # 检查文件大小（Excel文件通常至少有几千字节）
        if os.path.getsize(filepath) < 100:
            return False

        # 读取文件头部检查Excel文件标识
        with open(filepath, 'rb') as f:
            header = f.read(8)
            # Excel文件的魔法数字
            if header.startswith(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'):  # DOC/XLS 格式
                return True
            # 新版Excel格式
            if header.startswith(b'PK\x03\x04'):  # ZIP格式（包括XLSX）
                return True

        return False
    except Exception:
        return False


def excel_to_csv(excel_path: str, csv_path: str, sheet_name: Optional[str] = None) -> bool:
    """
    将Excel文件转换为CSV文件，处理合并单元格
    
    Args:
        excel_path: Excel文件路径
        csv_path: 输出的CSV文件路径
        sheet_name: 工作表名称，如果为None则取第一个工作表
        
    Returns:
        bool: 是否成功转换
    """
    try:
        print(f"正在转换Excel文件: {excel_path}")
        
        # 首先检查是否是HTML文件（优先于文件扩展名检查）
        if _is_html_table_file(excel_path):
            return _convert_html_table_to_csv(excel_path, csv_path)
        
        # 根据文件扩展名选择合适的库
        file_ext = os.path.splitext(excel_path)[1].lower()
        
        if file_ext == '.xlsx':
            # 使用openpyxl读取.xlsx文件
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            # 选择工作表
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    print(f"错误: 工作表 '{sheet_name}' 不存在")
                    return False
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title
            
            print(f"处理工作表: {sheet_name}")
            
            # 获取所有合并单元格
            merged_cells = ws.merged_cells.ranges
            
            # 创建数据字典来存储合并单元格的值
            merged_values = {}
            for merged_range in merged_cells:
                # 获取合并区域的第一个单元格（左上角）
                top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                top_left_value = top_left_cell.value
                
                # 为合并区域内的所有单元格记录这个值
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_values[(row, col)] = top_left_value
            
            # 收集所有数据
            data = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                row_data = []
                for col_idx, cell_value in enumerate(row, 1):
                    # 如果是合并单元格，使用合并区域的值
                    if (row_idx, col_idx) in merged_values:
                        row_data.append(merged_values[(row_idx, col_idx)])
                    else:
                        row_data.append(cell_value)
                data.append(row_data)
        
        elif file_ext == '.xls':
            
            # 使用xlrd读取.xls文件
            import xlrd
            wb = xlrd.open_workbook(excel_path, formatting_info=True)
            
            # 选择工作表
            if sheet_name:
                try:
                    ws = wb.sheet_by_name(sheet_name)
                except xlrd.XLRDError:
                    print(f"错误: 工作表 '{sheet_name}' 不存在")
                    return False
            else:
                ws = wb.sheet_by_index(0)
                sheet_name = ws.name
            
            print(f"处理工作表: {sheet_name}")
            
            # 获取所有合并单元格
            merged_cells = ws.merged_cells
            
            # 创建数据字典来存储合并单元格的值
            merged_values = {}
            for rlo, rhi, clo, chi in merged_cells:
                # 获取合并区域的第一个单元格（左上角）
                top_left_value = ws.cell_value(rlo, clo)
                
                # 为合并区域内的所有单元格记录这个值
                for row in range(rlo, rhi):
                    for col in range(clo, chi):
                        merged_values[(row, col)] = top_left_value
            
            # 收集所有数据
            data = []
            for row_idx in range(ws.nrows):
                row_data = []
                for col_idx in range(ws.ncols):
                    # 如果是合并单元格，使用合并区域的值
                    if (row_idx, col_idx) in merged_values:
                        row_data.append(merged_values[(row_idx, col_idx)])
                    else:
                        cell_value = ws.cell_value(row_idx, col_idx)
                        row_data.append(cell_value)
                data.append(row_data)
        
        else:
            print(f"错误: 不支持的文件格式: {file_ext}")
            return False
        
        # 转换为DataFrame并保存为CSV
        df = pd.DataFrame(data)
        df.to_csv(csv_path, index=False, header=False, encoding='utf-8-sig')
        
        print(f"成功转换: {excel_path} -> {csv_path}")
        print(f"处理了 {len(merged_cells)} 个合并单元格")
        return True
        
    except Exception as e:
        print(f"转换失败: {e}")
        return False



